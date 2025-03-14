import os, re, copy, json
from openpyxl import Workbook, load_workbook


def show_tab_overview(save_root):

    def find_column_index(sheet, column_name):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=1, column=col).value == column_name:
                return col
        return None


    accp_path = os.path.join(save_root, 'accp.xlsx')
    nqa_path = os.path.join(save_root, 'nqa.xlsx')
    final_path = os.path.join(save_root, 'final.xlsx')
    output_path = os.path.join(save_root, 'tab_overview.xlsx')


    accp_wb = load_workbook(accp_path)
    nqa_wb = load_workbook(nqa_path)
    final_wb = load_workbook(final_path)
    accp_sheet = accp_wb.active
    nqa_sheet = nqa_wb.active
    final_sheet = final_wb.active

    def find_column_index(sheet, column_name):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=1, column=col).value == column_name:
                return col
        return None

    accp_regular_col = find_column_index(accp_sheet, "regular")
    accp_extra_col = find_column_index(accp_sheet, "extra")
    nqa_regular_col = find_column_index(nqa_sheet, "regular")
    nqa_extra_col = find_column_index(nqa_sheet, "extra")
    final_all_col = find_column_index(final_sheet, "all")

    wb = Workbook()
    ws = wb.active
    headers = ['method', 'Reg_accp', 'Reg_nqa', 'Reg_avg', 'Ext_accp', 'Ext_nqa', 'Ext_avg', 'Avg']
    ws.append(headers)

    for row in range(2, min(len(accp_sheet['A']), len(nqa_sheet['A']), len(final_sheet['A'])) + 1):

        method = accp_sheet.cell(row=row, column=1).value

        reg_accp = accp_sheet.cell(row=row, column=accp_regular_col).value
        reg_nqa = nqa_sheet.cell(row=row, column=nqa_regular_col).value

        if reg_accp is not None and reg_nqa is not None:
            reg_avg = float(reg_accp) * 0.8 + float(reg_nqa) * 0.2
        else:
            reg_avg = None

        ext_accp = accp_sheet.cell(row=row, column=accp_extra_col).value
        ext_nqa = nqa_sheet.cell(row=row, column=nqa_extra_col).value

        if ext_accp is not None and ext_nqa is not None:
            ext_avg = float(ext_accp) * 0.8 + float(ext_nqa) * 0.2
        else:
            ext_avg = None

        avg = final_sheet.cell(row=row, column=final_all_col).value
        ws.append([method, reg_accp, reg_nqa, reg_avg, ext_accp, ext_nqa, ext_avg, avg])

    wb.save(output_path)

def show_tab_charttype_metric(save_root, typ='acc'):

    def find_column_index(sheet, column_name):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=1, column=col).value == column_name:
                return col
        return None

    accp_path = os.path.join(save_root, f'{typ}.xlsx')
    output_path = os.path.join(save_root, f'tab_charttype_{typ}.xlsx')

    accp_wb = load_workbook(accp_path)
    accp_sheet = accp_wb.active
    accp_line_col = find_column_index(accp_sheet, "line")
    accp_bar_col = find_column_index(accp_sheet, "bar")
    accp_pie_col = find_column_index(accp_sheet, "pie")
    accp_regular_col = find_column_index(accp_sheet, "regular")
    accp_area_col = find_column_index(accp_sheet, "area")
    accp_box_col = find_column_index(accp_sheet, "box")
    accp_radar_col = find_column_index(accp_sheet, "radar")
    accp_scatter_col = find_column_index(accp_sheet, "scatter")
    accp_node_col = find_column_index(accp_sheet, "node")
    accp_combin_col = find_column_index(accp_sheet, "combination")
    accp_extra_col = find_column_index(accp_sheet, "extra")
    accp_all_col = find_column_index(accp_sheet, "all")

    wb = Workbook()
    ws = wb.active

    headers = ['method', 'line', 'bar', 'pie', 'Reg_avg', 'area', 'box', 'radar','scatter', 'node', 'combination', 'Ext_avg', 'Avg']
    ws.append(headers)

    for row in range(2, len(accp_sheet['A']) + 1):
        method = accp_sheet.cell(row=row, column=1).value

        line = float(accp_sheet.cell(row=row, column=accp_line_col).value)
        bar = float(accp_sheet.cell(row=row, column=accp_bar_col).value)
        pie = float(accp_sheet.cell(row=row, column=accp_pie_col).value)
        Reg_avg = float(accp_sheet.cell(row=row, column=accp_regular_col).value)
        area = float(accp_sheet.cell(row=row, column=accp_area_col).value)
        box = float(accp_sheet.cell(row=row, column=accp_box_col).value)
        radar = float(accp_sheet.cell(row=row, column=accp_radar_col).value)
        scatter = float(accp_sheet.cell(row=row, column=accp_scatter_col).value)
        node = float(accp_sheet.cell(row=row, column=accp_node_col).value)
        combin = float(accp_sheet.cell(row=row, column=accp_combin_col).value)
        Ext_avg = float(accp_sheet.cell(row=row, column=accp_extra_col).value)
        avg = float(accp_sheet.cell(row=row, column=accp_all_col).value)

        ws.append([method, line, bar, pie, Reg_avg, area, box, radar, scatter, node, combin, Ext_avg, avg])

    wb.save(output_path)


def show_tab_charttype(save_root):

    def find_column_index(sheet, column_name):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=1, column=col).value == column_name:
                return col
        return None

    accp_path = os.path.join(save_root, 'accp.xlsx')
    nqa_path = os.path.join(save_root, 'nqa.xlsx')
    final_path = os.path.join(save_root, 'final.xlsx')
    output_path = os.path.join(save_root, 'tab_charttype.xlsx')

    accp_wb = load_workbook(accp_path)
    nqa_wb = load_workbook(nqa_path)
    final_wb = load_workbook(final_path)

    accp_sheet = accp_wb.active
    nqa_sheet = nqa_wb.active
    final_sheet = final_wb.active

    accp_line_col = find_column_index(accp_sheet, "line")
    accp_bar_col = find_column_index(accp_sheet, "bar")
    accp_pie_col = find_column_index(accp_sheet, "pie")
    accp_regular_col = find_column_index(accp_sheet, "regular")
    accp_area_col = find_column_index(accp_sheet, "area")
    accp_box_col = find_column_index(accp_sheet, "box")
    accp_radar_col = find_column_index(accp_sheet, "radar")
    accp_scatter_col = find_column_index(accp_sheet, "scatter")
    accp_node_col = find_column_index(accp_sheet, "node")
    accp_combin_col = find_column_index(accp_sheet, "combination")
    accp_extra_col = find_column_index(accp_sheet, "extra")

    nqa_line_col = find_column_index(nqa_sheet, "line")
    nqa_bar_col = find_column_index(nqa_sheet, "bar")
    nqa_pie_col = find_column_index(nqa_sheet, "pie")
    nqa_regular_col = find_column_index(nqa_sheet, "regular")
    nqa_area_col = find_column_index(nqa_sheet, "area")
    nqa_box_col = find_column_index(nqa_sheet, "box")
    nqa_radar_col = find_column_index(nqa_sheet, "radar")
    nqa_scatter_col = find_column_index(nqa_sheet, "scatter")
    nqa_node_col = find_column_index(nqa_sheet, "node")
    nqa_combin_col = find_column_index(nqa_sheet, "combination")
    nqa_extra_col = find_column_index(nqa_sheet, "extra")

    final_all_col = find_column_index(final_sheet, "all")

    wb = Workbook()
    ws = wb.active

    headers = ['method', 'line', 'bar', 'pie', 'Reg_avg', 'area', 'box', 'radar','scatter', 'node', 'combination', 'Ext_avg', 'Avg']
    ws.append(headers)

    for row in range(2, min(len(accp_sheet['A']), len(nqa_sheet['A']), len(final_sheet['A'])) + 1):
        method = accp_sheet.cell(row=row, column=1).value

        def get_weighted_value(accp_col, nqa_col):
            accp_val = accp_sheet.cell(row=row, column=accp_col).value
            nqa_val = nqa_sheet.cell(row=row, column=nqa_col).value
            try:
                accp_val = float(accp_val)
                nqa_val = float(nqa_val)
                return accp_val * 0.8 + nqa_val * 0.2
            except (ValueError, TypeError):
                return None

        line = get_weighted_value(accp_line_col, nqa_line_col)
        bar = get_weighted_value(accp_bar_col, nqa_bar_col)
        pie = get_weighted_value(accp_pie_col, nqa_pie_col)
        Reg_avg = get_weighted_value(accp_regular_col, nqa_regular_col)
        area = get_weighted_value(accp_area_col, nqa_area_col)
        box = get_weighted_value(accp_box_col, nqa_box_col)
        radar = get_weighted_value(accp_radar_col, nqa_radar_col)
        scatter = get_weighted_value(accp_scatter_col, nqa_scatter_col)
        node = get_weighted_value(accp_node_col, nqa_node_col)
        combin = get_weighted_value(accp_combin_col, nqa_combin_col)
        Ext_avg = get_weighted_value(accp_extra_col, nqa_extra_col)

        try:
            avg = float(final_sheet.cell(row=row, column=final_all_col).value)
        except (ValueError, TypeError):
            avg = None

        ws.append([method, line, bar, pie, Reg_avg, area, box, radar, scatter, node, combin, Ext_avg, avg])

    wb.save(output_path)


def show_tab_tasktype(save_root):

    def find_column_index(sheet, column_name):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=1, column=col).value == column_name:
                return col
        return None

    accp_path = os.path.join(save_root, 'accp.xlsx')
    cor_path = os.path.join(save_root, 'cor.xlsx')
    nqa_path = os.path.join(save_root, 'nqa.xlsx')
    final_path = os.path.join(save_root, 'final.xlsx')
    output_path = os.path.join(save_root, 'tab_tasktype.xlsx')

    accp_wb = load_workbook(accp_path)
    cor_wb = load_workbook(cor_path)
    nqa_wb = load_workbook(nqa_path)
    final_wb = load_workbook(final_path)

    accp_sheet = accp_wb.active
    cor_sheet = cor_wb.active
    nqa_sheet = nqa_wb.active
    final_sheet = final_wb.active

    cr_acc_col = find_column_index(accp_sheet, "CR")
    cr_cor_col = find_column_index(cor_sheet, "CR")
    ve_acc_col = find_column_index(accp_sheet, "VE")
    ve_cor_col = find_column_index(cor_sheet, "VE")
    vc_acc_col = find_column_index(accp_sheet, "VC")
    vc_cor_col = find_column_index(cor_sheet, "VC")
    gc_acc_col = find_column_index(accp_sheet, "GC")
    gc_cor_col = find_column_index(cor_sheet, "GC")
    nqa_col = find_column_index(nqa_sheet, "all")
    final_all_col = find_column_index(final_sheet, "all")

    wb = Workbook()
    ws = wb.active

    headers = ['method', 'cr_acc', 'cr_cor','ve_acc','ve_cor', 'vc_acc', 'vc_cor', 'gc_acc', 'gc_cor', 'nqa', 'avg']
    ws.append(headers)


    def get_value(sheet, col, row):
        try:
            return float(sheet.cell(row=row, column=col).value)
        except (ValueError, TypeError):
            return None

    for row in range(2, min(len(accp_sheet['A']), len(cor_sheet['A']), len(nqa_sheet['A']), len(final_sheet['A'])) + 1):
        method = accp_sheet.cell(row=row, column=1).value
        cr_acc = get_value(accp_sheet, cr_acc_col, row)
        cr_cor = get_value(cor_sheet, cr_cor_col, row)
        ve_acc = get_value(accp_sheet, ve_acc_col, row)
        ve_cor = get_value(cor_sheet, ve_cor_col, row)
        vc_acc = get_value(accp_sheet, vc_acc_col, row)
        vc_cor = get_value(cor_sheet, vc_cor_col, row)
        gc_acc = get_value(accp_sheet, gc_acc_col, row)
        gc_cor = get_value(cor_sheet, gc_cor_col, row)
        nqa = get_value(nqa_sheet, nqa_col, row)
        avg = get_value(final_sheet, final_all_col, row)

        ws.append([method, cr_acc, cr_cor, ve_acc, ve_cor, vc_acc, vc_cor, gc_acc, gc_cor, nqa, avg])

    wb.save(output_path)


def show_tab_anno(save_root):

    def find_column_index(sheet, column_name):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=1, column=col).value == column_name:
                return col
        return None

    accp_path = os.path.join(save_root, 'accp.xlsx')
    nqa_path = os.path.join(save_root, 'nqa.xlsx')
    output_path = os.path.join(save_root, 'tab_anno.xlsx')

    accp_wb = load_workbook(accp_path)
    nqa_wb = load_workbook(nqa_path)

    accp_sheet = accp_wb.active
    nqa_sheet = nqa_wb.active

    cr_wi_col = find_column_index(accp_sheet, "wi_CR")
    cr_wo_col = find_column_index(accp_sheet, "wo_CR")
    ve_wi_col = find_column_index(accp_sheet, "wi_VE")
    ve_wo_col = find_column_index(accp_sheet, "wo_VE")
    vc_wi_col = find_column_index(accp_sheet, "wi_VC")
    vc_wo_col = find_column_index(accp_sheet, "wo_VC")
    gc_wi_col = find_column_index(accp_sheet, "wi_GC")
    gc_wo_col = find_column_index(accp_sheet, "wo_GC")
    nqa_wi_col = find_column_index(nqa_sheet, "wi_anno")
    nqa_wo_col = find_column_index(nqa_sheet, "wo_anno")

    wb = Workbook()
    ws = wb.active

    headers = ['method', 'cr_wi', 'cr_wo','ve_wi','ve_wo', 'vc_wi', 'vc_wo', 'gc_wi', 'gc_wo', 'nqa_wi', 'nqa_wo',
            'avg_wi', 'avg_wo', 'delta']
    ws.append(headers)

    def get_value(sheet, col, row):
        try:
            return float(sheet.cell(row=row, column=col).value)
        except (ValueError, TypeError):
            return None

    for row in range(2, min(len(accp_sheet['A']), len(nqa_sheet['A'])) + 1):
        method = accp_sheet.cell(row=row, column=1).value
        cr_wi = get_value(accp_sheet, cr_wi_col, row)
        cr_wo = get_value(accp_sheet, cr_wo_col, row)
        ve_wi = get_value(accp_sheet, ve_wi_col, row)
        ve_wo = get_value(accp_sheet, ve_wo_col, row)
        vc_wi = get_value(accp_sheet, vc_wi_col, row)
        vc_wo = get_value(accp_sheet, vc_wo_col, row)
        gc_wi = get_value(accp_sheet, gc_wi_col, row)
        gc_wo = get_value(accp_sheet, gc_wo_col, row)
        nqa_wi = get_value(nqa_sheet, nqa_wi_col, row)
        nqa_wo = get_value(nqa_sheet, nqa_wo_col, row)

        values_wi = [val for val in [cr_wi, ve_wi, vc_wi, gc_wi, nqa_wi] if val is not None]
        avg_wi = sum(values_wi) / len(values_wi) if values_wi else None

        values_wo = [val for val in [cr_wo, ve_wo, vc_wo, gc_wo, nqa_wo] if val is not None]
        avg_wo = sum(values_wo) / len(values_wo) if values_wo else None

        delta = avg_wi - avg_wo if avg_wi is not None and avg_wo is not None else None

        ws.append([method, cr_wi, cr_wo, ve_wi, ve_wo, vc_wi, vc_wo, gc_wi, gc_wo, nqa_wi, nqa_wo, avg_wi, avg_wo, delta])

    wb.save(output_path)


def show_latex_pattern(save_root):
    patterns = {
        'tab_overview': r'\multicolumn{1}{l|}{METHOD} & #1# & #2# & \multicolumn{1}{c|}{#3#} & #4# & #5# & \multicolumn{1}{c|}{#6#} & #7# & \multicolumn{1}{c|}{-} & - & - & - & - \\',
        'tab_tasktype': r'\multicolumn{1}{l|}{METHOD} & #1# & \multicolumn{1}{c|}{#2#} & #3# & \multicolumn{1}{c|}{#4#} & #5# & \multicolumn{1}{c|}{#6#} & #7# & \multicolumn{1}{c|}{#8#} & \multicolumn{1}{c|}{#9#} & #10# \\',
        'tab_anno': r'\multicolumn{1}{l|}{METHOD} & #1# & \multicolumn{1}{c|}{#2#} & #3# & \multicolumn{1}{c|}{#4#} & #5# & \multicolumn{1}{c|}{#6#} & #7# & \multicolumn{1}{c|}{#8#} & #9# & \multicolumn{1}{c|}{#10#} & #11# & #12# & -#13# \\',
        'tab_charttype_acc': r'METHOD & #1# & #2# & #3# & #4# & #5# & #6# & #7# & #8# & #9# & #10# & #11# & #12# \\',
        'tab_charttype_accp': r'\multicolumn{1}{l|}{METHOD} & #1# & #2# & #3# & \multicolumn{1}{c|}{#4#} & #5# & #6# & #7# & #8# & #9# & #10# & \multicolumn{1}{c|}{#11#} & #12# \\',
        'tab_charttype_nqa': r'\multicolumn{1}{l|}{METHOD} & #1# & #2# & #3# & \multicolumn{1}{c|}{#4#} & #5# & #6# & #7# & #8# & #9# & #10# & \multicolumn{1}{c|}{#11#} & #12# \\',
        'tab_charttype_cor': r'\multicolumn{1}{l|}{METHOD} & #1# & #2# & #3# & \multicolumn{1}{c|}{#4#} & #5# & #6# & #7# & #8# & #9# & #10# & \multicolumn{1}{c|}{#11#} & #12# \\',
    }

    for item in patterns.keys():

        data_path = os.path.join(save_root, f'{item}.xlsx')
        save_path = os.path.join(save_root, f'latex_{item}.tex')
        pattern = patterns[item]

        wb = load_workbook(data_path)
        sheet = wb.active
        with open(save_path, 'w', encoding='utf-8') as file:
            
            for row in range(2, sheet.max_row + 1):
                values = []
                for col in range(1, sheet.max_column + 1):
                    cell_value = sheet.cell(row=row, column=col).value
                    if cell_value is not None:
                        values.append(str(cell_value))
                    else:
                        values.append('')
                method = values[0]
                rest_values = values[1:]
                line = pattern.replace("METHOD", method)
                for i, num in enumerate(rest_values):
                    line = line.replace(f"#{i+1}#", '{:.2f}'.format(float(num)))
                file.write(line + '\n')



if __name__ == '__main__':

    save_root = '../Result/Paper_Table'

    show_tab_overview(save_root)
    show_tab_charttype(save_root)
    show_tab_tasktype(save_root)
    show_tab_anno(save_root)
    show_tab_charttype_metric(save_root, 'accp')
    show_tab_charttype_metric(save_root, 'acc')
    show_tab_charttype_metric(save_root, 'nqa')
    show_tab_charttype_metric(save_root, 'cor')
    show_latex_pattern(save_root)

