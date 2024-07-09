import re, copy, json
from typing import Optional

metric_group = {
    'box': ['box_h', 'box_v', 'stock'],
    'combination': ['bar_line', 'line_line', 'pie_bar', 'pie_pie'],
    'pie': ['sector', 'ring_wo_anno', 'pie', 'ring', 'InteSun'],
    'scatter': ['scatter_2d', 'scatter_2d_smooth', 'scatter_3d'],
    'line': ['line_err', 'line_multi_wi_anno', 'line_multi', 'line_single_wi_anno', 'line_single'],
    'bar': ['horizontal_single', 'vertical_single', 'horizontal_single_wi_anno', 'vertical_single_wi_anno', 
            'vertical_percent_stacked', 'horizontal_multi', 'vertical_multi', 'threeD_stacked', 'vertical_stacked', 
            'horizontal_stacked', 'threeD_bar_multi', 'horizontal_percent_stacked', 'threeD_percent_stacked'],
    'radar': ['radar_single_wi_anno', 'radar_single', 'radar_multi_fill', 'radar_multi'],
    'area': ['area', 'area_stack', 'area_percent'],
    'node': ['node_link', 'node_link_dir', 'node_link_undir'],
    'wi_anno': ["horizontal_single_wi_anno", "vertical_single_wi_anno", "pie_pie", "pie_bar", 
                "radar_single_wi_anno", "node_link_dir", "node_link_undir", "ring_wi_anno", 
                "line_multi_wi_anno", "line_single_wi_anno"],
    'wo_anno': ["horizontal_single", "vertical_single", "bar_line", "line_line", "radar_single", 
                "ring", "line_multi", "line_single"]
}

metric_record_acc = {
    "all": [], "regular": [], "extra": [], "CR": [], "VE": [], "VC": [], "GC": [],
    "line": [], "bar": [], "pie": [], "area": [], "box": [], "radar": [], "scatter": [], "node": [], "combination": [],
    "wi_anno": [], "wo_anno": [], "wi_CR": [], "wo_CR": [], "wi_VE": [], "wo_VE": [], "wi_VC": [], "wo_VC": [], 
    "wi_GC": [], "wo_GC": []
}

metric_record_nqa = {
    "all": [], "regular": [], "extra": [], 
    "line": [], "bar": [], "pie": [], "area": [], "box": [], "radar": [], "scatter": [], "node": [], "combination": [],
    "wi_anno": [], "wo_anno": [], 
}

def relaxed_correctness(target: str,
                        prediction: str,
                        max_relative_change: float = 0.05) -> bool:

    def _prediction_to_float(text: str) -> Optional[float]:
        try:
            if text.endswith('%'):
                # Convert percentages to floats.
                # return float(text.rstrip('%')) / 100.0
                return float(text.rstrip('%'))
            else:
                return float(text)
        except ValueError:
            return None
        
    def _target_to_float(text: str) -> Optional[float]:
        try:
            if text.endswith('%'):
                # Convert percentages to floats.
                # return float(text.rstrip('%')) / 100.0
                return [float(text.rstrip('%')), float(text.rstrip('%')) / 100.0]
            else:
                return [float(text)]
        except ValueError:
            return None
        
    prediction_float = _prediction_to_float(prediction)
    target_float = _target_to_float(target)
    if prediction_float is not None and target_float is not None:
        flag = False
        # print(target_float)
        for t in target_float:
            if t == 0: 
                relative_change = prediction_float
            else:
                relative_change = abs(prediction_float - t) / abs(t)
            flag = flag or relative_change <= max_relative_change
        return flag
    else:
        return prediction.lower() == target.lower()


def fuzzy_match(sentence):
    sentence = str(sentence)
    contains_yes = re.search(r'[\[\{\(\s]*yes[\]\}\)\s\.,;]*', sentence, re.IGNORECASE) is not None
    contains_no = re.search(r'[\[\{\(\s]*no[\]\}\)\s\.,;]*', sentence, re.IGNORECASE) is not None
    contains_yes = 'yes' in sentence.lower()
    return contains_yes, not contains_yes
    # return not contains_no, contains_no
    # return contains_yes, contains_no

def error_plus(judgement):
    assert len(judgement) == 2
    _, isNo = fuzzy_match(judgement[0]["answer"])
    isYes, _ = fuzzy_match(judgement[1]["answer"])
    return isYes and isNo

def accuracy_plus(judgement):
    assert len(judgement) == 2
    isYes, _ = fuzzy_match(judgement[0]["answer"])
    _, isNo = fuzzy_match(judgement[1]["answer"])
    return isYes and isNo

def accuracy_vanilla(judgement):
    assert len(judgement) == 2
    isYes, _ = fuzzy_match(judgement[0]["answer"])
    _, isNo = fuzzy_match(judgement[1]["answer"])
    return [isYes, isNo]

def confuse_rate(judgement):
    assert len(judgement) == 2
    ar_yes, ar_no = fuzzy_match(judgement[0]["answer"])
    aw_yes, aw_no = fuzzy_match(judgement[1]["answer"])
    return (ar_yes and aw_yes) or (ar_no and aw_no)

def format_percent_metric(item):
    if len(item) == 0: return 0
    return sum(item) / len(item) * 100

def eval_all_metric_in_chartbench(results):
    def update_yes_no(metrics, key, accp, cor, acc, err):
        metrics['accp'][key].append(accp)
        metrics['cor'][key].append(cor)
        metrics['acc'][key].extend(acc)
        metrics['err'][key].append(err)
        
    def update_nqa(metrics, key, nqa):
        metrics['nqa'][key].append(nqa)

    def classify_item(item, metrics):
        chart_type = item["type"]["image"]
        task_type = item["type"]["task"]
        
        if item["type"]["QA"] == 'Acc+':
            accp = accuracy_plus(item["conversation"])
            cor = confuse_rate(item["conversation"])
            acc = accuracy_vanilla(item["conversation"])
            err = not accp and not cor
            update_yes_no(metrics, 'all', accp, cor, acc, err)

            for group_key, group_values in metric_group.items():
                if chart_type in group_values:
                    metric_category = 'regular' if group_key in {'line', 'bar', 'pie'} else 'extra'
                    update_yes_no(metrics, group_key, accp, cor, acc, err)
                    update_yes_no(metrics, metric_category, accp, cor, acc, err)

            anno_key = 'wi_anno' if chart_type in metric_group['wi_anno'] else 'wo_anno'
            update_yes_no(metrics, anno_key, accp, cor, acc, err)

            if task_type:
                update_yes_no(metrics, task_type, accp, cor, acc, err)
                task_anno_key = f'wi_{task_type}' if chart_type in metric_group['wi_anno'] else f'wo_{task_type}'
                update_yes_no(metrics, task_anno_key, accp, cor, acc, err)
                
        if item["type"]["QA"] == 'GPT-acc':
            if isinstance(item['conversation'][0]['label'], str):
                item['conversation'][0]['label'] = [item['conversation'][0]['label']]
            if 'gpt_filter' in item['conversation'][0].keys():
                eval_answer = item['conversation'][0]['gpt_filter']
            else:
                eval_answer = item['conversation'][0]['answer']
            nqa = max([
                    relaxed_correctness(eval_answer.strip(), ann)
                    for ann in item['conversation'][0]['label']
            ])
            update_nqa(metrics, 'all', nqa)

            for group_key, group_values in metric_group.items():
                if chart_type in group_values:
                    metric_category = 'regular' if group_key in {'line', 'bar', 'pie'} else 'extra'
                    update_nqa(metrics, group_key, nqa)
                    update_nqa(metrics, metric_category, nqa)

            anno_key = 'wi_anno' if chart_type in metric_group['wi_anno'] else 'wo_anno'
            update_nqa(metrics, anno_key, nqa)

    metrics = {
        'accp': copy.deepcopy(metric_record_acc),
        'cor': copy.deepcopy(metric_record_acc),
        'acc': copy.deepcopy(metric_record_acc),
        'err': copy.deepcopy(metric_record_acc),
        'nqa': copy.deepcopy(metric_record_nqa),
    }

    for item in results: classify_item(item, metrics)

    ans_stat = {key: {k: format_percent_metric(v) for k, v in metrics[key].items()} for key in metrics}
    return ans_stat

def eval_one_model(result_path):
    model_name = result_path.split('/')[-1].replace('.jsonl', '')
    with open(result_path, 'r') as fp:
        lines = fp.readlines()
        answers = [json.loads(l) for l in lines]
    results = eval_all_metric_in_chartbench(answers)
    return [results, model_name]

def show_final_table(stat_paths):
    from prettytable import PrettyTable
    from openpyxl import Workbook
    
    results = [eval_one_model(p) for p in stat_paths]
    print(results)
    
    for metric in ['acc', 'cor', 'accp', 'err', 'nqa']:
        table = PrettyTable()
        table.field_names = [metric] + list(results[0][0][metric].keys())
        for res in results:
            row = [res[1]] + [f'{i:.2f}' for i in res[0][metric].values()]
            table.add_row(row)
        print(table)

        workbook = Workbook()
        worksheet = workbook.active
        for col, header in enumerate(table.field_names, start=1):
            worksheet.cell(row=1, column=col, value=header)
        for i, row in enumerate(table.rows, start=2):
            for j, cell in enumerate(row, start=1):
                worksheet.cell(row=i, column=j, value=cell)

        workbook.save(f"./Paper_Table/{metric}.xlsx")


if __name__ == '__main__':
    stat_paths = ['/data/FinAi_Mapping_Knowledge/qiyiyan/xzz/ChartLLM/ChartBench/Result/raw/InternLM-XComposer-v2.jsonl']
    show_final_table(stat_paths)
